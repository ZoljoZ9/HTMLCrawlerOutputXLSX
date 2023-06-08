import os
import glob
import openpyxl

from bs4 import BeautifulSoup
import codecs

path = "C:\\Users\\matthewz\\Downloads\\lksd\\"

titleList = []
streetAddressList = []

for infile in glob.glob(os.path.join(path, "*.html")):
    markup = infile
    with codecs.open(markup, "r", encoding="utf-8") as file:
        content = file.read()
    soup = BeautifulSoup(content, 'lxml')
    title = soup.title.string if soup.title else ""
    address_tag = soup.find("address", class_="styles_address__zrPvy")
    address = address_tag.text if address_tag else ""
    titleList.append(title)
    streetAddressList.append(address)

# Create a new workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write the data to the worksheet
for i, (title, address) in enumerate(zip(titleList, streetAddressList), start=1):
    sheet.cell(row=i, column=1, value=title)
    sheet.cell(row=i, column=2, value=address)

# Save the workbook as XLSX file
output_file = "C:\\Users\\matthewz\\Downloads\\output2.xlsx"
workbook.save(output_file)
